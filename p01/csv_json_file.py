"""
题目：
1.拿到平安银行一年的股票数据 csv文件
2.里面一共244个交易日，我们读取csv文件，然后找到成交量(amount)大于1百万手的交易日的数据

3.然后把大于1百万手的那天的如下数据:
open
high
low
close
amount
写入excel文件
"""
import csv
import openpyxl

headers= ['ts_code', 'open', 'high', 'low', 'close', 'amount']


def read_csv(filename='000001.csv', max_amount=1000000):
    with open(filename) as file:
        result = []
        reader = csv.reader(file)
        for row in reader:
            bank_stock = {}
            if reader.line_num == 1:
                continue
            if float(row[-1]) > max_amount:
                bank_stock['ts_code'] = row[1]
                bank_stock['open'] = row[3]
                bank_stock['high'] = row[4]
                bank_stock['low'] = row[5]
                bank_stock['close'] = row[6]
                bank_stock['amount'] = row[-1]
                result.append(bank_stock)
    return result


def write_into_excel(data=[]):

    wb = openpyxl.Workbook()
    wb.create_sheet(index=0, title='stock')
    sheet = wb.active
    for i in range(len(headers)):
        sheet.cell(row=1, column=i+1).value = headers[i]

    for i, stock in enumerate(data):
        temp_list = [value for value in stock.values()]
        for j in range(len(temp_list)):
            sheet.cell(row=2 + i, column=1 + j).value = temp_list[j]

    wb.save(r'example.xlsx')


def main():
    data = read_csv()
    write_into_excel(data)


if __name__ == '__main__':
    main()
